import openpyxl

workbookPath = '../TransmissionStorage.xlsx'
workbook = openpyxl.load_workbook(workbookPath)
sheet = workbook['MainSheet']

rawValueRow = 11
rawMaxColumn = 2

maxERow = 1
maxARow = 2
maxAccRow = 3
maxVelRow = 4
maxForceRow = 5
maxPressureRow = 6
maxTempRow = 7

maxAyRow = 14
maxAxRow = 15
maxAzRow = 16

maxVyRow = 19
maxVxRow = 20
maxVzRow = 21

maxFyRow = 24
maxFxRow = 25
maxFzRow = 26

rawTimeColumn = 8
rawModeColumn = 10

rawAxColumn = 12
rawAyColumn = 13
rawAzColumn = 14
rawA_AbsoulteColumn = 15

rawVxColumn = 17
rawVyColumn = 18
rawVzColumn = 19
rawV_AbsoluteColumn = 20

rawHeightEnvColumn = 22
rawHeightAccColumn = 23

rawFxColumn = 25
rawFyColumn = 26
rawFzColumn = 27
rawF_AbsoluteColumn = 28

rawPressureColumn = 30
rawTemperatureColumn =31

rawXTiltColumn = 33
rawZTiltColumn = 34
rawAbsoluteTiltColumn = 35

def appendExcel(time, modeText, accelX, accelY, accelZ, accelVertical, velX, velY, velZ, velVertical, forceX, forceY, forceZ, forceVertical, altitudeE, altitudeA, pressure, temperature, maxAccelVert, aYMax, aXMax, aZMax, maxVelVert, vXMax, vYMax, vZMax, maxForceVert, fYMax, fXMax, fZMax, maxAltE, maxAltA, maxPressure, maxTemperature, launchTime, descendingTime, landedTime, avgTransmissionTime, Xtilt, Ztilt, absoluteTilt, minAccelVert, minForceVert, minVelVert, minAltA, minAltE, aXMin, aYMin, aZMin, vXMin, vYMin, vZMin, fXMin, fYMin, fZMin, XtiltMax, XtiltMin, ZtiltMax, ZtiltMin):
    sheet.cell(row=rawValueRow+len(time), column=rawModeColumn).value=modeText[-1]
    sheet.cell(row=rawValueRow+len(time), column=rawAxColumn).value=accelX[-1]
    sheet.cell(row=rawValueRow+len(time), column=rawAyColumn).value=accelY[-1]
    sheet.cell(row=rawValueRow+len(time), column=rawAzColumn).value=accelZ[-1]
    sheet.cell(row=rawValueRow+len(time), column=rawA_AbsoulteColumn).value=accelVertical[-1]
    sheet.cell(row=rawValueRow+len(time), column=rawVxColumn).value=velX[-1]
    sheet.cell(row=rawValueRow+len(time), column=rawVyColumn).value=velY[-1]
    sheet.cell(row=rawValueRow+len(time), column=rawVzColumn).value=velZ[-1]
    sheet.cell(row=rawValueRow+len(time), column=rawHeightEnvColumn).value=altitudeE[-1]
    sheet.cell(row=rawValueRow+len(time), column=rawHeightAccColumn).value=altitudeA[-1]
    sheet.cell(row=rawValueRow+len(time), column=rawFxColumn).value=forceX[-1]
    sheet.cell(row=rawValueRow+len(time), column=rawFyColumn).value=forceY[-1]
    sheet.cell(row=rawValueRow+len(time), column=rawFzColumn).value=forceZ[-1]
    sheet.cell(row=rawValueRow+len(time), column=rawF_AbsoluteColumn).value=forceVertical[-1]
    sheet.cell(row=rawValueRow+len(time), column=rawPressureColumn).value=pressure[-1]
    sheet.cell(row=rawValueRow+len(time), column=rawTemperatureColumn).value=temperature[-1]
    sheet.cell(row=rawValueRow+len(time), column=rawXTiltColumn).value=Xtilt[-1]
    sheet.cell(row=rawValueRow+len(time), column=rawZTiltColumn).value=Ztilt[-1]
    sheet.cell(row=rawValueRow+len(time), column=rawAbsoluteTiltColumn).value=absoluteTilt[-1]

    sheet.cell(row=maxERow, column = rawMaxColumn).value = maxAltE
    sheet.cell(row=maxARow, column = rawMaxColumn).value = maxAltA
    sheet.cell(row=maxAccRow, column = rawMaxColumn).value = maxAccelVert
    sheet.cell(row=maxVelRow, column = rawMaxColumn).value = maxVelVert
    sheet.cell(row=maxForceRow, column = rawMaxColumn).value = maxForceVert
    sheet.cell(row=maxPressureRow, column = rawMaxColumn).value = maxPressure
    sheet.cell(row=maxTempRow, column = rawMaxColumn).value = maxTemperature

    sheet.cell(row=maxAyRow, column = rawMaxColumn).value = aYMax
    sheet.cell(row=maxAxRow, column = rawMaxColumn).value = aXMax
    sheet.cell(row=maxAzRow, column = rawMaxColumn).value = aZMax

    sheet.cell(row=maxVyRow, column = rawMaxColumn).value = vYMax
    sheet.cell(row=maxVxRow, column = rawMaxColumn).value = vXMax
    sheet.cell(row=maxVzRow, column = rawMaxColumn).value = vZMax

    sheet.cell(row=maxFyRow, column = rawMaxColumn).value = fYMax
    sheet.cell(row=maxFxRow, column = rawMaxColumn).value = fXMax
    sheet.cell(row=maxFzRow, column = rawMaxColumn).value = fZMax
    sheet['D3']=launchTime #Stores the time of launch
    sheet['E3']=descendingTime #Stores when descent begins
    sheet['F3']=landedTime #Stores the time of landing
    sheet['I3']=avgTransmissionTime #Stores the time of landing
    workbook.save(filename="/Users/eoghancollins/ArduinoLocal/TransmissionStorageReal.xlsx")



# def appendMaxs(accelMax, ayMax, axMax, azMax, velMax, vyMax, vxMax, vzMax, forceMax, fyMax, fxMax, fzMax, heightEmax, heightAmax, pressureMax, temperatureMax):
#     sheet.cell(row=maxERow, column = rawMaxColumn).value = heightEmax
#     sheet.cell(row=maxARow, column = rawMaxColumn).value = heightAmax
#     sheet.cell(row=maxAccRow, column = rawMaxColumn).value = accelMax
#     sheet.cell(row=maxVelRow, column = rawMaxColumn).value = velMax
#     sheet.cell(row=maxForceRow, column = rawMaxColumn).value = forceMax
#     sheet.cell(row=maxPressureRow, column = rawMaxColumn).value = pressureMax
#     sheet.cell(row=maxTempRow, column = rawMaxColumn).value = temperatureMax

#     sheet.cell(row=maxAyRow, column = rawMaxColumn).value = ayMax
#     sheet.cell(row=maxAxRow, column = rawMaxColumn).value = axMax
#     sheet.cell(row=maxAzRow, column = rawMaxColumn).value = azMax

#     sheet.cell(row=maxVyRow, column = rawMaxColumn).value = vyMax
#     sheet.cell(row=maxVxRow, column = rawMaxColumn).value = vxMax
#     sheet.cell(row=maxVzRow, column = rawMaxColumn).value = vzMax

#     sheet.cell(row=maxFyRow, column = rawMaxColumn).value = fyMax
#     sheet.cell(row=maxFxRow, column = rawMaxColumn).value = fxMax
#     sheet.cell(row=maxFzRow, column = rawMaxColumn).value = fzMax
#     workbook.save(filename="/Users/eoghancollins/ArduinoLocal/TransmissionStorageReal.xlsx")


# # def appendModeTime(mode, time):


        
